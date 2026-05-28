import json
import boto3
import urllib.request
import urllib.parse
import io
from datetime import datetime

bedrock = boto3.client(service_name='bedrock-runtime', region_name='ap-northeast-2')
secrets_client = boto3.client(service_name='secretsmanager', region_name='ap-northeast-2')

QUESTIONS_FOLDER_ID = '1FKFvIV560vOgNE7hQih8mmvBOKeRAAZq'
QUOTE_FOLDER_ID = '17Pgodymgr-Kze51aPsUCKOm7GtPzdCQE'

AWS_PRICING = {
    'EC2': {
        't3.micro':   {'ondemand': 0.0116, 'ri': 0.0072, 'vcpu': 2,  'ram': 1},
        't3.small':   {'ondemand': 0.0232, 'ri': 0.0144, 'vcpu': 2,  'ram': 2},
        't3.medium':  {'ondemand': 0.0576, 'ri': 0.0347, 'vcpu': 2,  'ram': 4},
        't3.large':   {'ondemand': 0.1152, 'ri': 0.0694, 'vcpu': 2,  'ram': 8},
        't3.xlarge':  {'ondemand': 0.2304, 'ri': 0.1386, 'vcpu': 4,  'ram': 16},
        'm5.large':   {'ondemand': 0.124,  'ri': 0.075,  'vcpu': 2,  'ram': 8},
        'm5.xlarge':  {'ondemand': 0.248,  'ri': 0.149,  'vcpu': 4,  'ram': 16},
        'm5.2xlarge': {'ondemand': 0.496,  'ri': 0.298,  'vcpu': 8,  'ram': 32},
        'c5.large':   {'ondemand': 0.107,  'ri': 0.064,  'vcpu': 2,  'ram': 4},
        'c5.xlarge':  {'ondemand': 0.214,  'ri': 0.128,  'vcpu': 4,  'ram': 8},
    },
    'RDS': {
        'db.t3.micro':   {'ondemand': 0.034, 'ri': 0.021},
        'db.t3.small':   {'ondemand': 0.068, 'ri': 0.041},
        'db.t3.medium':  {'ondemand': 0.136, 'ri': 0.082},
        'db.t3.large':   {'ondemand': 0.272, 'ri': 0.163},
        'db.r6g.large':  {'ondemand': 0.288, 'ri': 0.173},
        'db.r6g.xlarge': {'ondemand': 0.576, 'ri': 0.346},
    },
    'EBS':        {'gp3': 0.0912},
    'S3':         {'standard': 0.025},
    'ALB':        {'hourly': 0.0225},
    'NAT':        {'hourly': 0.059},
    'WAF':        {'per_acl': 5.0},
    'CloudFront': {'per_gb': 0.114},
}


def calculate_pricing(services, exchange_rate):
    for svc in services:
        name = svc.get('service_name', '')
        qty = svc.get('qty', 1)
        instance_type = svc.get('type', '')

        if 'EC2' in name:
            pricing = AWS_PRICING['EC2'].get(instance_type, {})
            svc['ondemand_hourly'] = pricing.get('ondemand', 0)
            svc['ri_hourly'] = pricing.get('ri', 0)
            svc['ondemand_usd'] = round(qty * svc['ondemand_hourly'] * 730, 3)
            svc['ri_usd'] = round(qty * svc['ri_hourly'] * 730, 3)
            if not svc.get('vcpu'):
                svc['vcpu'] = pricing.get('vcpu', '')
            if not svc.get('ram'):
                svc['ram'] = pricing.get('ram', '')
        elif 'RDS' in name:
            pricing = AWS_PRICING['RDS'].get(instance_type, {})
            svc['ondemand_hourly'] = pricing.get('ondemand', 0)
            svc['ri_hourly'] = pricing.get('ri', 0)
            svc['ondemand_usd'] = round(qty * svc['ondemand_hourly'] * 730, 3)
            svc['ri_usd'] = round(qty * svc['ri_hourly'] * 730, 3)
        elif 'EBS' in name:
            size = svc.get('size_gb', 0)
            svc['ondemand_usd'] = round(size * AWS_PRICING['EBS']['gp3'], 3)
            svc['ri_usd'] = svc['ondemand_usd']
        elif 'S3' in name:
            size = svc.get('size_gb', 0)
            svc['ondemand_usd'] = round(size * AWS_PRICING['S3']['standard'], 3)
            svc['ri_usd'] = svc['ondemand_usd']
        elif 'ALB' in name or 'ELB' in name:
            svc['ondemand_usd'] = round(AWS_PRICING['ALB']['hourly'] * 730, 3)
            svc['ri_usd'] = svc['ondemand_usd']
        elif 'NAT' in name:
            svc['ondemand_usd'] = round(AWS_PRICING['NAT']['hourly'] * 730, 3)
            svc['ri_usd'] = svc['ondemand_usd']
        elif 'WAF' in name:
            svc['ondemand_usd'] = AWS_PRICING['WAF']['per_acl'] * qty
            svc['ri_usd'] = svc['ondemand_usd']
        elif 'CloudFront' in name:
            size = svc.get('size_gb', 0)
            svc['ondemand_usd'] = round(size * AWS_PRICING['CloudFront']['per_gb'], 3)
            svc['ri_usd'] = svc['ondemand_usd']

    return services


def get_secret():
    secret = secrets_client.get_secret_value(SecretId='sales-agent/google-credentials')
    return json.loads(secret['SecretString'])


def get_access_token(secret):
    data = urllib.parse.urlencode({
        'client_id': secret['client_id'],
        'client_secret': secret['client_secret'],
        'refresh_token': secret['refresh_token'],
        'grant_type': 'refresh_token'
    }).encode()
    req = urllib.request.Request('https://oauth2.googleapis.com/token', data=data, method='POST')
    with urllib.request.urlopen(req) as response:
        return json.loads(response.read())['access_token']


def create_questions_doc(questions_text, customer_name, access_token):
    today = datetime.now().strftime("%Y%m%d")
    doc_title = f"[질문지] {customer_name}_{today}"

    data = json.dumps({
        'name': doc_title,
        'mimeType': 'application/vnd.google-apps.document',
        'parents': [QUESTIONS_FOLDER_ID]
    }).encode()

    req = urllib.request.Request(
        'https://www.googleapis.com/drive/v3/files',
        data=data,
        headers={'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'},
        method='POST'
    )
    try:
        with urllib.request.urlopen(req) as response:
            doc = json.loads(response.read())
    except urllib.error.HTTPError as e:
        print(f"DOC CREATE ERROR: {e.read().decode()}")
        raise

    doc_id = doc['id']

    data = json.dumps({
        'requests': [{'insertText': {'location': {'index': 1}, 'text': questions_text}}]
    }).encode()

    req = urllib.request.Request(
        f'https://docs.googleapis.com/v1/documents/{doc_id}:batchUpdate',
        data=data,
        headers={'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'},
        method='POST'
    )
    try:
        with urllib.request.urlopen(req) as response:
            response.read()
    except urllib.error.HTTPError as e:
        print(f"DOC INSERT ERROR: {e.read().decode()}")
        raise

    return doc_id, f"https://docs.google.com/document/d/{doc_id}/edit"


def create_quote_xlsx(quote_data):
    """템플릿 없이 직접 가견적서 엑셀을 생성한다."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = openpyxl.Workbook()

    # ── Total 시트 ──────────────────────────────────────────────────
    ws_total = wb.active
    ws_total.title = 'Total'

    today = datetime.now().strftime("%Y-%m-%d")
    customer_name = quote_data.get('customer_name', '고객사')
    sales_rep = quote_data.get('sales_rep', '담당자 미정')
    exchange_rate = quote_data.get('exchange_rate', 1450)
    services = quote_data.get('services', [])

    # 고정 정보
    ws_total['B2'] = 'AWS'
    ws_total['B4'] = '고객사'
    ws_total['D4'] = customer_name
    ws_total['F4'] = '공급사'
    ws_total['H4'] = '메가존클라우드㈜'
    ws_total['B5'] = '경기도 과천시 과천대로7길 74, 메가존산학연센터'
    ws_total['B6'] = '부서/담당자'
    ws_total['D6'] = '-'
    ws_total['F6'] = '사업자등록번호 / 대표이사'
    ws_total['H6'] = '232-88-00982 / 이주완'
    ws_total['B7'] = '작성일 (견적유효기간)'
    ws_total['D7'] = today
    ws_total['F7'] = '(견적 일로부터 1개월 이내)'
    ws_total['H7'] = '견적 및 영업담당'
    ws_total['I7'] = sales_rep

    # 비용 계산
    total_ondemand_usd = sum(s.get('ondemand_usd', 0) for s in services)
    total_ri_usd = sum(s.get('ri_usd', 0) for s in services)
    total_ondemand_krw = round(total_ondemand_usd * exchange_rate)
    total_ri_krw = round(total_ri_usd * exchange_rate)

    ws_total['B9'] = '견적 제안'
    ws_total['B11'] = 'AWS Infra 비용'
    ws_total['D11'] = f'₩ {total_ondemand_krw:,}'
    ws_total['H11'] = f'(VAT별도) 1. 환율은 {exchange_rate:,}원으로 계산 / 2. 온디맨드 결제옵션 기준'

    ws_total['B13'] = 'AWS infra 약정 비용(1y)'
    ws_total['D13'] = f'₩ {total_ri_krw:,}'
    ws_total['H13'] = f'(VAT별도) 1. 환율은 {exchange_rate:,}원으로 계산 / 2. 선납 없음 기준'

    ws_total['B15'] = '전체 비용'
    ws_total['D15'] = f'₩ {total_ondemand_krw:,}'
    ws_total['H15'] = '(VAT별도)'

    ws_total['B16'] = '전체 비용 (VAT 포함)'
    ws_total['D16'] = f'₩ {round(total_ondemand_krw * 1.1):,}'
    ws_total['H16'] = '(VAT포함)'

    ws_total['B18'] = '* 참고사항'
    ws_total['B19'] = '1. 본 예상 비용 산출서는 AWS 월 이용 요금에 대한 예상 금액을 산출한 견적서입니다.'
    ws_total['B20'] = '2. 실제 청구되는 월 AWS 비용은 사용 환경 및 사용량에 따라 상이 할 수 있습니다.'
    ws_total['B21'] = '3. 고객사에서 보내 주신 정보를 토대로 작성하였으며 사양 및 옵션의 조정이 필요할 수 있습니다.'
    ws_total['B22'] = "4. 결제 옵션 중 RI는 모두 '선결제없음'기준이며, Convertible RI 기준입니다."

    # ── 상세내역 시트 ──────────────────────────────────────────────
    ws_detail = wb.create_sheet('상세내역')

    # 헤더
    ws_detail['B2'] = 'AWS Infra 상세 내역'
    ws_detail['B4'] = '1. 매월 납입비 (Amazon Web Services - Singapore Region)'

    # 테이블 헤더 (Row 7)
    headers = ['AWS\n서비스 명', 'Server Name', 'OS', 'Qty', 'Type',
               'vCPU', 'RAM\n(GiB)', 'Size\n(GB)', 'On-Demand\n(USD/월)',
               'RI 1y\n(USD/월)', 'On-Demand\n(KRW/월)', 'RI\n(KRW/월)', '비고']
    for col, h in enumerate(headers, start=2):
        cell = ws_detail.cell(row=7, column=col, value=h)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 환율 표시
    ws_detail['N5'] = f'환율: ₩{exchange_rate:,}'

    # 서비스별 데이터 입력
    ec2_services = [s for s in services if 'EC2' in s.get('service_name', '')]
    rds_services = [s for s in services if 'RDS' in s.get('service_name', '')]
    ebs_services = [s for s in services if 'EBS' in s.get('service_name', '')]
    s3_services  = [s for s in services if 'S3' in s.get('service_name', '')]
    alb_services = [s for s in services if 'ALB' in s.get('service_name', '') or 'ELB' in s.get('service_name', '')]
    nat_services = [s for s in services if 'NAT' in s.get('service_name', '')]
    waf_services = [s for s in services if 'WAF' in s.get('service_name', '')]

    row = 9  # 데이터 시작 행

    # EC2
    for svc in ec2_services:
        ws_detail.cell(row=row, column=2, value='EC2\n(Amazon EC2)')
        ws_detail.cell(row=row, column=3, value=svc.get('server_name', ''))
        ws_detail.cell(row=row, column=4, value=svc.get('os', 'Linux'))
        ws_detail.cell(row=row, column=5, value=svc.get('qty', 1))
        ws_detail.cell(row=row, column=6, value=svc.get('type', ''))
        ws_detail.cell(row=row, column=7, value=svc.get('vcpu', ''))
        ws_detail.cell(row=row, column=8, value=svc.get('ram', ''))
        ws_detail.cell(row=row, column=9, value=svc.get('size_gb', ''))
        ws_detail.cell(row=row, column=10, value=round(svc.get('ondemand_usd', 0), 2))
        ws_detail.cell(row=row, column=11, value=round(svc.get('ri_usd', 0), 2))
        ws_detail.cell(row=row, column=12, value=round(svc.get('ondemand_usd', 0) * exchange_rate))
        ws_detail.cell(row=row, column=13, value=round(svc.get('ri_usd', 0) * exchange_rate))
        ws_detail.cell(row=row, column=14, value=svc.get('note', ''))
        row += 1

    # RDS
    for svc in rds_services:
        ws_detail.cell(row=row, column=2, value='RDS\n(Amazon RDS)')
        ws_detail.cell(row=row, column=3, value=svc.get('server_name', 'DB Server'))
        ws_detail.cell(row=row, column=4, value=svc.get('os', 'RDS for MySQL'))
        ws_detail.cell(row=row, column=5, value=svc.get('qty', 1))
        ws_detail.cell(row=row, column=6, value=svc.get('type', ''))
        ws_detail.cell(row=row, column=7, value=svc.get('vcpu', ''))
        ws_detail.cell(row=row, column=8, value=svc.get('ram', ''))
        ws_detail.cell(row=row, column=9, value=svc.get('size_gb', ''))
        ws_detail.cell(row=row, column=10, value=round(svc.get('ondemand_usd', 0), 2))
        ws_detail.cell(row=row, column=11, value=round(svc.get('ri_usd', 0), 2))
        ws_detail.cell(row=row, column=12, value=round(svc.get('ondemand_usd', 0) * exchange_rate))
        ws_detail.cell(row=row, column=13, value=round(svc.get('ri_usd', 0) * exchange_rate))
        ws_detail.cell(row=row, column=14, value=svc.get('note', ''))
        row += 1

    # EBS
    for svc in ebs_services:
        ws_detail.cell(row=row, column=2, value='EBS\n(Amazon EBS)')
        ws_detail.cell(row=row, column=5, value=svc.get('size_gb', 0))
        ws_detail.cell(row=row, column=10, value=round(svc.get('ondemand_usd', 0), 2))
        ws_detail.cell(row=row, column=11, value=round(svc.get('ri_usd', 0), 2))
        ws_detail.cell(row=row, column=12, value=round(svc.get('ondemand_usd', 0) * exchange_rate))
        ws_detail.cell(row=row, column=13, value=round(svc.get('ri_usd', 0) * exchange_rate))
        ws_detail.cell(row=row, column=14, value=svc.get('note', 'GB당 $0.0912'))
        row += 1

    # S3
    for svc in s3_services:
        ws_detail.cell(row=row, column=2, value='S3\n(Amazon S3)')
        ws_detail.cell(row=row, column=5, value=svc.get('size_gb', 0))
        ws_detail.cell(row=row, column=10, value=round(svc.get('ondemand_usd', 0), 2))
        ws_detail.cell(row=row, column=11, value=round(svc.get('ri_usd', 0), 2))
        ws_detail.cell(row=row, column=12, value=round(svc.get('ondemand_usd', 0) * exchange_rate))
        ws_detail.cell(row=row, column=13, value=round(svc.get('ri_usd', 0) * exchange_rate))
        ws_detail.cell(row=row, column=14, value=svc.get('note', 'GB당 $0.025'))
        row += 1

    # WAF
    for svc in waf_services:
        ws_detail.cell(row=row, column=2, value='WAF\n(AWS WAF)')
        ws_detail.cell(row=row, column=5, value=svc.get('qty', 1))
        ws_detail.cell(row=row, column=10, value=round(svc.get('ondemand_usd', 0), 2))
        ws_detail.cell(row=row, column=11, value=round(svc.get('ri_usd', 0), 2))
        ws_detail.cell(row=row, column=12, value=round(svc.get('ondemand_usd', 0) * exchange_rate))
        ws_detail.cell(row=row, column=13, value=round(svc.get('ri_usd', 0) * exchange_rate))
        ws_detail.cell(row=row, column=14, value='웹 ACL당 $5')
        row += 1

    # ALB
    for svc in alb_services:
        ws_detail.cell(row=row, column=2, value='ALB\n(Elastic LB)')
        ws_detail.cell(row=row, column=5, value=svc.get('qty', 1))
        ws_detail.cell(row=row, column=10, value=round(svc.get('ondemand_usd', 0), 2))
        ws_detail.cell(row=row, column=11, value=round(svc.get('ri_usd', 0), 2))
        ws_detail.cell(row=row, column=12, value=round(svc.get('ondemand_usd', 0) * exchange_rate))
        ws_detail.cell(row=row, column=13, value=round(svc.get('ri_usd', 0) * exchange_rate))
        ws_detail.cell(row=row, column=14, value='시간당 $0.0225 × 730h')
        row += 1

    # NAT
    for svc in nat_services:
        ws_detail.cell(row=row, column=2, value='NAT Gateway')
        ws_detail.cell(row=row, column=5, value=svc.get('qty', 1))
        ws_detail.cell(row=row, column=10, value=round(svc.get('ondemand_usd', 0), 2))
        ws_detail.cell(row=row, column=11, value=round(svc.get('ri_usd', 0), 2))
        ws_detail.cell(row=row, column=12, value=round(svc.get('ondemand_usd', 0) * exchange_rate))
        ws_detail.cell(row=row, column=13, value=round(svc.get('ri_usd', 0) * exchange_rate))
        ws_detail.cell(row=row, column=14, value='시간당 $0.059 × 730h')
        row += 1

    # 합계 행
    row += 1
    ws_detail.cell(row=row, column=2, value='AWS Usage 합계 (USD) / 매월')
    ws_detail.cell(row=row, column=10, value=round(total_ondemand_usd, 2))
    ws_detail.cell(row=row, column=11, value=round(total_ri_usd, 2))
    ws_detail.cell(row=row, column=12, value=total_ondemand_krw)
    ws_detail.cell(row=row, column=13, value=total_ri_krw)
    ws_detail.cell(row=row, column=14, value='순수 AWS Infra 비용')
    ws_detail.cell(row=row, column=2).font = Font(bold=True)

    row += 1
    ws_detail.cell(row=row, column=2, value='합계')
    ws_detail.cell(row=row, column=6, value=f'₩ {exchange_rate:,}')
    ws_detail.cell(row=row, column=12, value=f'₩ {total_ondemand_krw:,}')
    ws_detail.cell(row=row, column=13, value=f'₩ {total_ri_krw:,}')
    ws_detail.cell(row=row, column=14, value='총 합계')
    ws_detail.cell(row=row, column=2).font = Font(bold=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def upload_xlsx_to_drive(xlsx_bytes, filename, access_token):
    boundary = 'boundary_sales_agent'
    metadata = json.dumps({
        'name': filename,
        'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'parents': [QUOTE_FOLDER_ID]
    })

    body = (
        f'--{boundary}\r\n'
        f'Content-Type: application/json; charset=UTF-8\r\n\r\n'
        f'{metadata}\r\n'
        f'--{boundary}\r\n'
        f'Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n'
    ).encode() + xlsx_bytes + f'\r\n--{boundary}--\r\n'.encode()

    req = urllib.request.Request(
        'https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart',
        data=body,
        headers={
            'Authorization': f'Bearer {access_token}',
            'Content-Type': f'multipart/related; boundary={boundary}'
        },
        method='POST'
    )
    try:
        with urllib.request.urlopen(req) as response:
            result = json.loads(response.read())
            file_id = result['id']
            return file_id, f'https://drive.google.com/file/d/{file_id}/view'
    except urllib.error.HTTPError as e:
        print(f"DRIVE UPLOAD ERROR: {e.read().decode()}")
        raise


def analyze_email(conversation):
    prompt = f"""당신은 AWS 클라우드 전문 영업 상담 에이전트입니다.

고객사는 대기업, 중소기업, 공공기관, 금융기관 등 다양하며
주요 서비스는 AWS 클라우드 인프라 구축 및 운영입니다.

아래는 영업자와 고객사 간의 메일 또는 대화 내용입니다.
전체 내용을 분석하여 아직 확인되지 않은 정보를 파악하고,
영업자가 고객과 추가 상담 시 반드시 확인해야 할 질문 리스트를 생성해주세요.

## 분석 기준

**[서비스 범위]**
- 어떤 AWS 서비스가 필요한지 명확하지 않은 경우
- 온프레미스에서 클라우드 전환인지, 신규 구축인지 불명확한 경우
- 멀티 클라우드 또는 하이브리드 환경 여부가 불명확한 경우

**[규모 및 트래픽]**
- 예상 사용자 수 또는 트래픽 규모 미언급
- 데이터 용량 및 스토리지 요구사항 불명확
- 가용성 요구사항(예: 99.9% SLA) 미언급

**[보안 및 컴플라이언스]**
- 보안 요구사항 미언급
- 개인정보보호, ISMS, CC인증 등 컴플라이언스 요건 불명확
- 네트워크 보안 요구사항 미언급

**[일정 및 예산]**
- 프로젝트 시작일 및 완료 목표일 미언급
- 예산 범위 미언급
- 단계적 도입 계획 여부 불명확

**[의사결정 구조]**
- 기술 담당자와 구매 결재자가 불명확한 경우
- 현재 사용 중인 클라우드 또는 인프라 환경 미언급

## 중요 지침
- 이미 대화에서 언급되거나 확인된 내용은 질문에서 반드시 제외
- 구어체, 대화체, 이메일 형식 등 어떤 형태의 입력이든 핵심 정보 추출
- 불명확하거나 누락된 항목만 질문으로 생성

## 출력 형식
---
### 📋 고객 상담 질문지

**[확인된 정보 요약]**
- ✅ [대화에서 이미 확인된 내용]

**[서비스 범위]**
1. 질문: [질문 내용]
   의도: [이 질문을 통해 확인해야 하는 것]

**[규모 및 트래픽]**
2. 질문: [질문 내용]
   의도: [이 질문을 통해 확인해야 하는 것]
...
---

대화 내용:
{conversation}"""

    response = bedrock.invoke_model(
        modelId='global.anthropic.claude-sonnet-4-6',
        body=json.dumps({
            "anthropic_version": "bedrock-2023-05-31",
            "max_tokens": 8192,
            "messages": [{"role": "user", "content": prompt}]
        })
    )
    return json.loads(response['body'].read())['content'][0]['text']


def generate_quote(customer_info):
    prompt = f"""당신은 AWS 클라우드 전문 영업 상담 에이전트입니다.

아래 상담 내용에서 필요한 AWS 서비스 정보를 JSON으로 추출해주세요.
요금 계산은 하지 않아도 됩니다. 서비스 종류, 수량, 스펙, 간단한 비고만 추출해주세요.

## 출력 형식 (JSON만 출력, 다른 텍스트 없이)
{{
    "customer_name": "고객사명",
    "sales_rep": "담당자 미정",
    "exchange_rate": 1450,
    "services": [
        {{
            "service_name": "EC2 (Amazon EC2)",
            "server_name": "WAS",
            "os": "Linux",
            "qty": 2,
            "type": "t3.large",
            "vcpu": 2,
            "ram": 8,
            "size_gb": 50,
            "note": "WAS 서버 2대"
        }}
    ]
}}

## 주의사항
- 언급되지 않은 서비스는 포함하지 마세요
- EC2 타입: t3.micro/t3.small/t3.medium/t3.large/t3.xlarge/m5.large/m5.xlarge/c5.large 중 선택
- RDS 타입: db.t3.micro/db.t3.small/db.t3.medium/db.t3.large/db.r6g.large 중 선택
- 사용자 수, 트래픽 규모에 따라 적절한 인스턴스 타입 선택
- note는 한 줄로 간결하게

상담 내용:
{customer_info}"""

    response = bedrock.invoke_model(
        modelId='global.anthropic.claude-sonnet-4-6',
        body=json.dumps({
            "anthropic_version": "bedrock-2023-05-31",
            "max_tokens": 2000,
            "messages": [{"role": "user", "content": prompt}]
        })
    )
    raw = json.loads(response['body'].read())['content'][0]['text']
    raw = raw.strip()
    if raw.startswith('```'):
        raw = raw.split('\n', 1)[1].rsplit('```', 1)[0]
    quote_data = json.loads(raw)
    exchange_rate = quote_data.get('exchange_rate', 1450)
    quote_data['services'] = calculate_pricing(quote_data['services'], exchange_rate)
    return quote_data


def lambda_handler(event, context):
    import traceback
    try:
        action_group = event.get('actionGroup', '')
        function_name = event.get('function', '')
        parameters = event.get('parameters', [])
        params = {p['name']: p['value'] for p in parameters}

        action = function_name if function_name else event.get('action', '')
        if not action:
            return {'statusCode': 400, 'body': json.dumps({'error': 'Invalid request'}, ensure_ascii=False)}

        result = {}

        if action == 'analyze_email':
            conversation = params.get('email_content') or event.get('email_content', '')
            customer_name = params.get('customer_name') or event.get('customer_name', '고객사')
            if not conversation:
                result = {'error': '대화 내용이 없습니다.'}
            else:
                questions = analyze_email(conversation)
                secret = get_secret()
                access_token = get_access_token(secret)
                doc_id, doc_url = create_questions_doc(questions, customer_name, access_token)
                result = {'questions': questions, 'doc_url': doc_url, 'doc_id': doc_id}

        elif action == 'generate_quote':
            customer_info = params.get('customer_info') or event.get('customer_info', '')
            sales_rep = params.get('sales_rep') or event.get('sales_rep', '담당자 미정')
            if not customer_info:
                result = {'error': '상담 내용이 없습니다.'}
            else:
                quote_data = generate_quote(customer_info)
                quote_data['sales_rep'] = sales_rep
                secret = get_secret()
                access_token = get_access_token(secret)
                xlsx_bytes = create_quote_xlsx(quote_data)
                today = datetime.now().strftime("%Y%m%d")
                customer_name = quote_data.get('customer_name', '고객사')
                filename = f"[가견적서] {customer_name}_{today}.xlsx"
                file_id, file_url = upload_xlsx_to_drive(xlsx_bytes, filename, access_token)
                result = {'quote_data': quote_data, 'file_url': file_url, 'file_id': file_id}

        else:
            result = {'error': f'알 수 없는 action: {action}'}

        if function_name:
            return {
                "messageVersion": "1.0",
                "response": {
                    "actionGroup": action_group,
                    "function": function_name,
                    "functionResponse": {
                        "responseBody": {
                            "TEXT": {"body": json.dumps(result, ensure_ascii=False)}
                        }
                    }
                }
            }
        else:
            return {
                'statusCode': 200,
                'body': json.dumps(result, ensure_ascii=False)
            }

    except Exception as e:
        error_msg = {'error': str(e), 'trace': traceback.format_exc()}
        if event.get('function'):
            return {
                "messageVersion": "1.0",
                "response": {
                    "actionGroup": event.get('actionGroup', ''),
                    "function": event.get('function', ''),
                    "functionResponse": {
                        "responseBody": {
                            "TEXT": {"body": json.dumps(error_msg, ensure_ascii=False)}
                        }
                    }
                }
            }
        else:
            return {
                'statusCode': 500,
                'body': json.dumps(error_msg, ensure_ascii=False)
            }
